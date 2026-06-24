import csv
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from sqlalchemy.orm import Session

from app.db.models import CollectionRule, IndexConstituent, Stock, StockCollectionSetting
from app.domains.collection import repository
from app.domains.collection.schemas import (
    CollectionRuleCreate,
    CollectionRuleUpdate,
    CollectionStockRead,
    CollectionStockSummary,
    CollectionStockUpdate,
    ImportResult,
    IndexConstituentImportRequest,
    RecalculateResult,
)


def _normalize_code(value: Any) -> str:
    return str(value).strip().replace(".0", "").zfill(6)


def _rows_from_matrix(rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []

    header_keys = {"stock_code", "code", "종목코드", "단축코드", "stock_name", "name", "종목명", "한글종목명"}
    header_index = 0
    for idx, row in enumerate(rows[:10]):
        normalized_headers = {str(cell).strip().lower() for cell in row if cell is not None}
        if normalized_headers & {key.lower() for key in header_keys}:
            header_index = idx
            break

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[header_index]]
    return [
        dict(zip(headers, row))
        for row in rows[header_index + 1 :]
        if any(cell is not None and str(cell).strip() != "" for cell in row)
    ]


def _read_rows(file_path: str) -> list[dict[str, Any]]:
    path = Path(file_path)
    if not path.exists():
        raise HTTPException(status_code=404, detail="import file not found")

    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(status_code=500, detail="openpyxl is required for Excel import") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        return _rows_from_matrix([list(row) for row in ws.iter_rows(values_only=True)])
    if suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise HTTPException(status_code=500, detail="xlrd is required for XLS import") from exc
        wb = xlrd.open_workbook(str(path))
        sheet = wb.sheet_by_index(0)
        if sheet.nrows == 0:
            return []
        return _rows_from_matrix([
            [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]
            for row_idx in range(sheet.nrows)
        ])

    raise HTTPException(status_code=400, detail="only csv/xls/xlsx/xlsm files are supported")


def _pick(row: dict[str, Any], keys: list[str]) -> Any:
    normalized = {str(k).strip().lower(): v for k, v in row.items()}
    for key in keys:
        value = normalized.get(key.lower())
        if value is not None and str(value).strip() != "":
            return value
    return None


def import_index_constituents(db: Session, payload: IndexConstituentImportRequest) -> ImportResult:
    rows = _read_rows(payload.file_path)
    now = datetime.utcnow()
    created_stock_count = 0
    updated_existing_count = 0

    previous = repository.list_index_constituents(db, index_code=payload.index_code, is_active=True)
    for item in previous:
        item.is_active = False
    deactivated_previous_count = len(previous)

    imported_count = 0
    for row in rows:
        code_value = _pick(row, ["stock_code", "code", "종목코드", "단축코드"])
        name_value = _pick(row, ["stock_name", "name", "종목명", "한글종목명"])
        if code_value is None or name_value is None:
            continue

        stock_code = _normalize_code(code_value)
        stock_name = str(name_value).strip()
        market = _pick(row, ["market", "시장", "시장구분"])
        market_value = str(market).strip() if market is not None else None

        stock = repository.get_stock_by_code(db, stock_code)
        if not stock:
            stock = Stock(code=stock_code, name=stock_name, market=market_value, is_active=True)
            db.add(stock)
            db.flush()
            created_stock_count += 1
        else:
            stock.name = stock_name
            if market_value:
                stock.market = market_value
            stock.is_active = True
            updated_existing_count += 1

        existing = db.query(IndexConstituent).filter(
            IndexConstituent.index_code == payload.index_code,
            IndexConstituent.stock_code == stock_code,
            IndexConstituent.effective_date == payload.effective_date,
        ).first()
        if existing:
            existing.index_name = payload.index_name
            existing.tracking_index = payload.tracking_index
            existing.stock_id = stock.id
            existing.stock_name = stock_name
            existing.market = market_value
            existing.source = payload.source
            existing.is_active = True
            existing.updated_at = now
        else:
            db.add(IndexConstituent(
                index_code=payload.index_code,
                index_name=payload.index_name,
                tracking_index=payload.tracking_index,
                stock_id=stock.id,
                stock_code=stock_code,
                stock_name=stock_name,
                market=market_value,
                effective_date=payload.effective_date,
                is_active=True,
                source=payload.source,
            ))
        imported_count += 1

    db.commit()
    return ImportResult(
        imported_count=imported_count,
        created_stock_count=created_stock_count,
        updated_existing_count=updated_existing_count,
        deactivated_previous_count=deactivated_previous_count,
    )


def get_index_constituents(db: Session, index_code: str | None = None):
    return repository.list_index_constituents(db, index_code=index_code, is_active=True)


def get_index_constituents_summary(db: Session):
    active = repository.list_index_constituents(db, is_active=True)
    return {
        "total_count": len(repository.list_index_constituents(db, is_active=None)),
        "active_count": len(active),
        "kodex_200_count": sum(1 for item in active if item.index_code == "KODEX_200"),
        "kodex_kosdaq150_count": sum(
            1 for item in active if item.index_code in {"KODEX_KOSDAQ_150", "KODEX_KOSDAQ150"}
        ),
    }


def _setting_or_default(stock: Stock, setting: StockCollectionSetting | None, is_holding: bool) -> CollectionStockRead:
    return CollectionStockRead(
        stock_id=stock.id,
        stock_code=stock.code,
        stock_name=stock.name,
        market=stock.market,
        sector=stock.sector,
        market_cap=stock.market_cap,
        current_price=stock.current_price,
        is_favorite=stock.is_favorite,
        is_holding_calculated=bool(is_holding),
        collect_enabled=setting.collect_enabled if setting else False,
        collect_news=setting.collect_news if setting else False,
        collect_price_snapshot=setting.collect_price_snapshot if setting else False,
        collect_alert_enabled=setting.collect_alert_enabled if setting else False,
        priority=setting.priority if setting else "low",
        collect_reason=setting.collect_reason if setting else None,
        manual_override=setting.manual_override if setting else False,
        manual_include=setting.manual_include if setting else False,
        manual_exclude=setting.manual_exclude if setting else False,
        last_collected_at=setting.last_collected_at if setting else None,
    )


def get_collection_stocks(db: Session, **filters):
    rows = repository.list_collection_stocks(db, **filters)
    return [_setting_or_default(stock, setting, is_holding) for stock, setting, is_holding in rows]


def get_collection_stocks_summary(db: Session) -> CollectionStockSummary:
    total, enabled, news, alert, manual_include, manual_exclude = repository.collection_summary_counts(db)
    return CollectionStockSummary(
        total_candidate_count=total,
        collect_enabled_count=enabled,
        collect_news_count=news,
        collect_alert_enabled_count=alert,
        manual_include_count=manual_include,
        manual_exclude_count=manual_exclude,
    )


def _get_or_create_setting(db: Session, stock_id: int) -> StockCollectionSetting:
    setting = repository.get_stock_collection_setting(db, stock_id)
    if setting:
        return setting
    setting = StockCollectionSetting(stock_id=stock_id)
    db.add(setting)
    db.flush()
    return setting


def update_collection_stock(db: Session, stock_id: int, payload: CollectionStockUpdate):
    setting = _get_or_create_setting(db, stock_id)
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(setting, key, value)
    if payload.manual_include is not None or payload.manual_exclude is not None:
        setting.manual_override = True
    db.commit()
    return get_collection_stock_detail(db, stock_id)


def get_collection_stock_detail(db: Session, stock_id: int):
    rows = repository.list_collection_stocks(db, keyword=None)
    for stock, setting, is_holding in rows:
        if stock.id == stock_id:
            return _setting_or_default(stock, setting, is_holding)
    raise HTTPException(status_code=404, detail="stock not found")


def include_collection_stock(db: Session, stock_id: int):
    return update_collection_stock(db, stock_id, CollectionStockUpdate(
        collect_enabled=True,
        collect_news=True,
        collect_price_snapshot=True,
        priority="high",
        collect_reason="manual_include",
        manual_override=True,
        manual_include=True,
        manual_exclude=False,
    ))


def exclude_collection_stock(db: Session, stock_id: int):
    return update_collection_stock(db, stock_id, CollectionStockUpdate(
        collect_enabled=False,
        collect_news=False,
        collect_price_snapshot=False,
        collect_alert_enabled=False,
        priority="low",
        collect_reason="manual_exclude",
        manual_override=True,
        manual_include=False,
        manual_exclude=True,
    ))


def _matches_rule(stock: Stock, index_codes: set[str], rule: CollectionRule) -> str | None:
    condition = rule.condition_json or {}
    if rule.rule_type == "index":
        targets = set(condition.get("target_indices") or [])
        if condition.get("index_code"):
            targets.add(condition["index_code"])
        if targets and index_codes & targets:
            return "index_rule"
    if rule.rule_type == "index_member":
        index_code = condition.get("index_code")
        if index_code and index_code in index_codes:
            return "index_rule"
    if rule.rule_type == "market_cap":
        minimum = condition.get("market_cap_min")
        if minimum is not None and stock.market_cap is not None and stock.market_cap >= minimum:
            return "market_cap_rule"
    if rule.rule_type == "market":
        markets = set(condition.get("markets") or [])
        if markets and stock.market in markets:
            return "index_rule"
    if rule.rule_type == "sector":
        sectors = set(condition.get("sectors") or [])
        if sectors and stock.sector in sectors:
            return "index_rule"
    return None


def recalculate_collection_stocks(db: Session) -> RecalculateResult:
    stocks = repository.list_active_stocks(db)
    rules = repository.list_active_rules(db)
    index_codes_by_stock = repository.get_active_index_codes_by_stock(db)
    holding_stock_ids = repository.get_holding_stock_ids(db)
    alert_stock_ids = repository.get_alert_stock_ids(db)
    collect_enabled_count = 0
    manual_exclude_count = 0

    for stock in stocks:
        setting = _get_or_create_setting(db, stock.id)
        reason = None
        collect_enabled = False
        priority = "low"

        if setting.manual_exclude:
            reason = "manual_exclude"
            collect_enabled = False
            priority = "low"
            manual_exclude_count += 1
        elif setting.manual_include:
            reason = "manual_include"
            collect_enabled = True
            priority = "high"
        elif stock.id in holding_stock_ids:
            reason = "holding"
            collect_enabled = True
            priority = "high"
        elif stock.is_favorite:
            reason = "favorite"
            collect_enabled = True
            priority = "high"
        elif stock.id in alert_stock_ids:
            reason = "alert"
            collect_enabled = True
            priority = "high"
        else:
            index_codes = index_codes_by_stock.get(stock.id, set())
            for rule in rules:
                matched_reason = _matches_rule(stock, index_codes, rule)
                if matched_reason:
                    reason = matched_reason
                    collect_enabled = True
                    priority = "normal"
                    break

        setting.collect_enabled = collect_enabled
        setting.collect_news = collect_enabled
        setting.collect_price_snapshot = collect_enabled
        setting.collect_alert_enabled = stock.id in alert_stock_ids
        setting.priority = priority
        setting.collect_reason = reason
        if collect_enabled:
            collect_enabled_count += 1

    db.commit()
    return RecalculateResult(
        processed_count=len(stocks),
        collect_enabled_count=collect_enabled_count,
        manual_exclude_count=manual_exclude_count,
    )


def get_collection_rules(db: Session):
    return repository.list_collection_rules(db)


def create_collection_rule(db: Session, payload: CollectionRuleCreate):
    item = CollectionRule(**payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


def update_collection_rule(db: Session, rule_id: int, payload: CollectionRuleUpdate):
    item = repository.get_collection_rule(db, rule_id)
    if not item:
        raise HTTPException(status_code=404, detail="collection rule not found")
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(item, key, value)
    db.commit()
    db.refresh(item)
    return item


def delete_collection_rule(db: Session, rule_id: int) -> None:
    item = repository.get_collection_rule(db, rule_id)
    if not item:
        raise HTTPException(status_code=404, detail="collection rule not found")
    db.delete(item)
    db.commit()
